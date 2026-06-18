#!/usr/bin/env python3
"""Build JLCPCB/LCSC BOM review files from the KiCad schematic BOM export.

The KiCad export remains the source of truth for designators, values, and
assigned footprints. LCSC part numbers are resolved in priority order:
  1. Raw schematic LCSC column (set via KiCad symbol properties + sync script)
  2. bom/parts_db.csv keyed by (value, footprint)  ← preferred for new parts
  3. OVERRIDES dict below (deprecated — kept until schematic sync is complete)
  4. Empty string (unresolved)
"""

from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any


RAW_BOM = Path("bom/smart_pouch_kicad_bom_raw.csv")
TECHNICAL_WORKBOOK = Path("bom/smart_pouch_technical_reference.xlsx")
JLC_CART_WORKBOOK = Path("bom/JLC_CART.xlsx")
DEFAULT_PARTS_DB = Path("bom/parts_db.csv")


@dataclass(frozen=True)
class Override:
    lcsc: str = ""
    manufacturer: str = ""
    mpn: str = ""
    footprint: str = ""
    note: str = ""
    manual: bool = False


OVERRIDES: dict[str, Override] = {}  # deprecated — use bom/parts_db.csv

PART_WARNINGS: dict[str, str] = {}


def load_parts_db(path: Path) -> dict[tuple[str, str], dict[str, str]]:
    """Load parts_db.csv keyed by (value.lower(), footprint). Returns {} if missing."""
    if not path.exists():
        return {}
    db: dict[tuple[str, str], dict[str, str]] = {}
    with path.open(newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            key = (row["value"].lower().strip(), row["footprint"].strip())
            db[key] = row
    return db


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--raw", type=Path, default=RAW_BOM)
    parser.add_argument("--parts-db", type=Path, default=DEFAULT_PARTS_DB,
                        help="Path to parts_db.csv (default: bom/parts_db.csv)")
    parser.add_argument("--board-qty", type=int, default=2, help="Number of boards for cart/order quantities.")
    parser.add_argument("--no-network", action="store_true", help="Do not refresh JLCPCB data.")
    return parser.parse_args()


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def split_ref_token(token: str) -> tuple[str, int] | None:
    match = re.fullmatch(r"([A-Za-z]+)(\d+)", token.strip())
    if not match:
        return None
    return match.group(1), int(match.group(2))


def expand_refs(refs: str) -> list[str]:
    expanded: list[str] = []
    for item in refs.split(","):
        item = item.strip()
        if "-" not in item:
            expanded.append(item)
            continue
        start, end = item.split("-", 1)
        start_ref = split_ref_token(start)
        end_ref = split_ref_token(end)
        if not start_ref or not end_ref or start_ref[0] != end_ref[0]:
            expanded.append(item)
            continue
        prefix = start_ref[0]
        expanded.extend(f"{prefix}{number}" for number in range(start_ref[1], end_ref[1] + 1))
    return expanded


def dec(value: Any) -> Decimal | None:
    try:
        if value in (None, ""):
            return None
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def get_attr(obj: Any, name: str, default: Any = "") -> Any:
    if isinstance(obj, dict):
        return obj.get(name, default)
    return getattr(obj, name, default)


def normalize_search_results(results: Any) -> list[Any]:
    if isinstance(results, dict):
        inner = results.get("results", [])
        return inner if isinstance(inner, list) else []
    if isinstance(results, list):
        return results
    return list(results) if results else []


def fetch_live_parts(lcsc_ids: set[str], no_network: bool) -> dict[str, dict[str, Any]]:
    if no_network or not lcsc_ids:
        return {}
    try:
        from easyeda2kicad.easyeda.easyeda_api import EasyedaApi
    except Exception as exc:  # pragma: no cover - depends on local install
        print(f"Warning: easyeda2kicad API unavailable: {exc}")
        return {}

    api = EasyedaApi()
    live: dict[str, dict[str, Any]] = {}
    for lcsc_id in sorted(lcsc_ids):
        try:
            results = normalize_search_results(api.search_jlcpcb_components(lcsc_id, page_size=10))
        except Exception as exc:  # pragma: no cover - network dependent
            print(f"Warning: JLCPCB lookup failed for {lcsc_id}: {exc}")
            continue
        exact = None
        for result in results:
            if get_attr(result, "lcsc") == lcsc_id:
                exact = result
                break
        if exact is None and results:
            exact = results[0]
        if exact is None:
            continue
        live[lcsc_id] = {
            "lcsc": get_attr(exact, "lcsc"),
            "manufacturer": get_attr(exact, "brand"),
            "mpn": get_attr(exact, "model"),
            "package": get_attr(exact, "package"),
            "description": get_attr(exact, "description"),
            "library_type": get_attr(exact, "type"),
            "stock": get_attr(exact, "stock"),
            "min_qty": get_attr(exact, "min_qty"),
            "reel_qty": get_attr(exact, "reel_qty"),
            "price": get_attr(exact, "price"),
        }
    return live


def build_rows(raw_rows: list[dict[str, str]], live: dict[str, dict[str, Any]], board_qty: int, parts_db: dict[tuple[str, str], dict[str, str]] | None = None) -> list[dict[str, Any]]:
    if parts_db is None:
        parts_db = {}
    rows: list[dict[str, Any]] = []
    for raw in raw_rows:
        refs = raw.get("Refs") or raw.get("Reference", "")
        raw_value = raw.get("Value", "").strip()
        raw_footprint = raw.get("Footprint", "").strip()

        # Priority 1: LCSC from schematic (populated by sync_lcsc_to_sch.py)
        lcsc = raw.get("LCSC Part", raw.get("LCSC", "")).strip()

        # Priority 2: parts_db lookup by (value, footprint)
        db_entry: dict[str, str] = {}
        if not lcsc:
            db_entry = parts_db.get((raw_value.lower(), raw_footprint), {})
            lcsc = db_entry.get("lcsc", "").strip()

        # Priority 3: OVERRIDES (deprecated fallback)
        override = OVERRIDES.get(refs, Override())
        if not lcsc:
            lcsc = override.lcsc

        live_part = live.get(lcsc, {})

        # Footprint: parts_db > OVERRIDES > raw schematic
        effective_footprint = db_entry.get("footprint", "").strip() or override.footprint or raw_footprint
        qty_per_board = len(expand_refs(refs))
        order_qty = qty_per_board * board_qty
        min_qty = int(live_part.get("min_qty") or 0)
        suggested_cart_qty = max(order_qty, min_qty or order_qty)
        unit_price = dec(live_part.get("price"))
        order_total = unit_price * Decimal(order_qty) if unit_price is not None else None
        suggested_total = unit_price * Decimal(suggested_cart_qty) if unit_price is not None else None
        note_parts = []
        if db_entry.get("note"):
            note_parts.append(db_entry["note"])
        elif override.note:
            note_parts.append(override.note)
        if lcsc in PART_WARNINGS:
            note_parts.append(PART_WARNINGS[lcsc])
        if not lcsc:
            note_parts.append("No LCSC/JLCPCB part selected yet.")
        if override.manual:
            note_parts.append("Manual review required.")
        rows.append(
            {
                "Designators": ",".join(expand_refs(refs)),
                "Value": raw_value,
                "KiCad Footprint": effective_footprint,
                "LCSC Part #": lcsc,
                "Manufacturer": live_part.get("manufacturer") or db_entry.get("manufacturer") or override.manufacturer or raw.get("Manufacturer", ""),
                "MFR Part #": live_part.get("mpn") or db_entry.get("mpn") or override.mpn or raw.get("MPN", ""),
                "Description": live_part.get("description", ""),
                "Package": live_part.get("package", ""),
                "Library Type": live_part.get("library_type", ""),
                "Stock": live_part.get("stock", ""),
                "Min Qty": min_qty or "",
                "Reel Qty": live_part.get("reel_qty", ""),
                "Board Qty": board_qty,
                "Qty Per Board": qty_per_board,
                "Order Qty": order_qty,
                "Suggested Cart Qty": suggested_cart_qty if lcsc else "",
                "Unit Price": str(unit_price) if unit_price is not None else "",
                "Order Total": str(order_total) if order_total is not None else "",
                "Suggested Cart Total": str(suggested_total) if suggested_total is not None else "",
                "Notes": " ".join(note_parts),
            }
        )
    return rows



def _style_sheet(ws: Any, sheet_name: str) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(bottom=thin)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for data_row in ws.iter_rows(min_row=2):
        for cell in data_row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = max(
            (min(len("" if cell.value is None else str(cell.value)), 70) for cell in ws[letter]),
            default=0,
        )
        ws.column_dimensions[letter].width = max(10, min(max_len + 2, 55))

    if ws.max_row >= 2 and ws.max_column >= 2:
        table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        table = Table(displayName=f"{sheet_name.replace(' ', '')}Table", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)


def write_single_sheet_xlsx(path: Path, sheet_name: str, rows: list[dict[str, Any]], headers: list[str]) -> None:
    try:
        from openpyxl import Workbook
    except Exception as exc:  # pragma: no cover - depends on local install
        print(f"Warning: could not write {path}: openpyxl unavailable: {exc}")
        return

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    _style_sheet(ws, sheet_name)
    wb.save(path)


def main() -> None:
    args = parse_args()
    raw_rows = read_csv(args.raw)
    parts_db = load_parts_db(args.parts_db)

    # Collect LCSC IDs from all three priority sources for the live lookup.
    lcsc_ids: set[str] = set()
    for row in raw_rows:
        lcsc = row.get("LCSC Part", row.get("LCSC", "")).strip()
        if not lcsc:
            db_entry = parts_db.get((row.get("Value", "").lower().strip(), row.get("Footprint", "").strip()), {})
            lcsc = db_entry.get("lcsc", "").strip()
        if not lcsc:
            refs = row.get("Refs") or row.get("Reference", "")
            lcsc = OVERRIDES.get(refs, Override()).lcsc
        if lcsc:
            lcsc_ids.add(lcsc)

    live = fetch_live_parts(lcsc_ids, args.no_network)
    rows = build_rows(raw_rows, live, args.board_qty, parts_db)

    review_headers = [
        "Designators",
        "Value",
        "KiCad Footprint",
        "LCSC Part #",
        "Manufacturer",
        "MFR Part #",
        "Description",
        "Package",
        "Library Type",
        "Stock",
        "Min Qty",
        "Reel Qty",
        "Board Qty",
        "Qty Per Board",
        "Order Qty",
        "Suggested Cart Qty",
        "Unit Price",
        "Order Total",
        "Suggested Cart Total",
        "Notes",
    ]

    cart_headers = [
        "Parts Type", "JLCPCB Part #", "MFR Part #", "Description",
        "Unit Price", "Qty", "Total Price",
    ]
    cart_rows = [
        {
            "Parts Type": row["Library Type"],
            "JLCPCB Part #": row["LCSC Part #"],
            "MFR Part #": row["MFR Part #"],
            "Description": row["Description"],
            "Unit Price": row["Unit Price"],
            "Qty": row["Order Qty"],
            "Total Price": row["Order Total"],
        }
        for row in rows
        if row["LCSC Part #"]
    ]

    write_single_sheet_xlsx(JLC_CART_WORKBOOK, "Cart", cart_rows, cart_headers)
    write_single_sheet_xlsx(TECHNICAL_WORKBOOK, "Reference", rows, review_headers)

    print(f"Wrote {JLC_CART_WORKBOOK} ({len(cart_rows)} parts)")
    print(f"Wrote {TECHNICAL_WORKBOOK} ({len(rows)} parts)")


if __name__ == "__main__":
    main()
